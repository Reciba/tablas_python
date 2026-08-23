"""
Generador de archivos de muestra PDF y Excel con tablas y encabezados desplazados (filas de basura superiores)
para probar el sistema de extracción y limpieza.
"""

import os
import sys
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# Asegurar encoding utf-8 en Windows si es necesario
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def create_sample_pdf(output_path: str):
    """Crea un PDF de muestra con múltiples tablas y encabezados a partir de la fila 3 y 4."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    doc = SimpleDocTemplate(output_path, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    story = []
    styles = getSampleStyleSheet()

    # --- PÁGINA 1: TABLA 1 (Encabezados en la Fila 4, Filas 1-3 basura/metadatos) ---
    story.append(Paragraph("<b>DISTRIBUIDORA INDUSTRIAL S.A. - REPORTE DE VENTAS</b>", styles["Heading2"]))
    story.append(Spacer(1, 10))

    # Definición de la matriz de la Tabla 1
    # Fila 1: Título/Empresa
    # Fila 2: Metadato Factura / Fecha
    # Fila 3: Información de Cliente / Vendedor
    # Fila 4: ENCABEZADOS REALES
    # Filas 5-11: Datos
    # Fila 12: Pie / Notas
    tabla1_data = [
        ["DISTRIBUIDORA INDUSTRIAL S.A.", "", "", "", "", ""],                             # Fila 1 (Basura)
        ["Reporte No. 45892", "Fecha: 15/08/2026", "Sucursal: Centro", "", "Moneda: USD", ""],# Fila 2 (Basura)
        ["Vendedor: Carlos Morales", "Cliente: Consorcio Norte", "", "", "Estado: Cerrado", ""], # Fila 3 (Basura)
        ["Codigo", "Descripcion Producto", "Categoria", "Cantidad", "Precio Unitario", "Subtotal"], # Fila 4 (ENCABEZADOS REALES)
        ["PROD-101", "Sensor Ultrasonido Industrial", "Sensores", "15", "45.50", "682.50"],
        ["PROD-102", "Controlador PLC Modular", "Automatizacion", "4", "320.00", "1280.00"],
        ["PROD-103", "Valvula Solenoide 24V", "Neumatica", "22", "18.75", "412.50"],
        ["PROD-104", "Fuente de Poder 24V 10A", "Electrico", "8", "65.00", "520.00"],
        ["PROD-105", "Panel HMI Tactil 7 pulg", "Automatizacion", "2", "450.00", "900.00"],
        ["PROD-106", "Cable Blindado 4x1.5mm (100m)", "Cables", "3", "88.00", "264.00"],
        ["PROD-107", "Rele de Seguridad 2 Canales", "Seguridad", "6", "110.00", "660.00"],
        ["* NOTA: Precios netos antes de impuestos. Descuento comercial aplicado.", "", "", "", "", ""] # Fila 12 (Footer)
    ]

    t1 = Table(tabla1_data, colWidths=[65, 170, 95, 55, 80, 75])
    t1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 2), colors.whitesmoke), # Filas basura
        ('TEXTCOLOR', (0, 0), (-1, 2), colors.dimgrey),
        ('FONTNAME', (0, 0), (-1, 2), 'Helvetica-Oblique'),
        ('SPAN', (0, 0), (5, 0)),
        ('BACKGROUND', (0, 3), (-1, 3), colors.navy),       # Fila de encabezado real (Fila 4)
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.whitesmoke),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('ALIGN', (0, 3), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('SPAN', (0, 11), (5, 11)),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story.append(t1)

    story.append(PageBreak())

    # --- PÁGINA 2: TABLA 2 y TABLA 3 ---
    story.append(Paragraph("<b>DESGLOSE DE PAGOS Y LOGISTICA</b>", styles["Heading2"]))
    story.append(Spacer(1, 10))

    # Tabla 2: Encabezados en Fila 3
    tabla2_data = [
        ["RESUMEN FINANCIERO POR CENTRO DE COSTOS", "", "", ""],                # Fila 1 (Basura)
        ["Periodo Fiscal: Agosto 2026 | Responsable: Auditoria", "", "", ""],    # Fila 2 (Basura)
        ["Centro de Costos", "Presupuesto Asignado", "Ejecutado", "Disponible"], # Fila 3 (ENCABEZADO REAL)
        ["CC-100 Operaciones", "50000.00", "32450.00", "17550.00"],
        ["CC-200 Logistica", "28000.00", "19800.00", "8200.00"],
        ["CC-300 Mantencion", "15000.00", "11200.00", "3800.00"],
        ["CC-400 Administracion", "12000.00", "9400.00", "2600.00"]
    ]

    t2 = Table(tabla2_data, colWidths=[160, 120, 120, 120])
    t2.setStyle(TableStyle([
        ('SPAN', (0, 0), (3, 0)),
        ('SPAN', (0, 1), (3, 1)),
        ('BACKGROUND', (0, 2), (-1, 2), colors.darkslategray),
        ('TEXTCOLOR', (0, 2), (-1, 2), colors.whitesmoke),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story.append(t2)
    story.append(Spacer(1, 20))

    # Tabla 3: Encabezados en Fila 4
    tabla3_data = [
        ["DETALLE DE TRANSPORTE Y GUIAS DE DESPACHO", "", "", "", ""],          # Fila 1 (Basura)
        ["Transportista: Chilexpress / Starken", "", "", "", ""],               # Fila 2 (Basura)
        ["Zona: Metropolitana y Region de Valparaiso", "", "", "", ""],         # Fila 3 (Basura)
        ["No. Guia", "Destino", "Bultos", "Peso (Kg)", "Estado Entrega"],       # Fila 4 (ENCABEZADOS REALES)
        ["G-99810", "Santiago Centro", "3", "14.5", "Entregado"],
        ["G-99811", "Vina del Mar", "1", "2.8", "En Ruta"],
        ["G-99812", "Rancagua", "5", "42.0", "Entregado"],
        ["G-99813", "Quilicura", "2", "8.1", "Preparacion"],
        ["G-99814", "Valparaiso", "4", "19.3", "Entregado"]
    ]

    t3 = Table(tabla3_data, colWidths=[80, 150, 70, 90, 130])
    t3.setStyle(TableStyle([
        ('SPAN', (0, 0), (4, 0)),
        ('SPAN', (0, 1), (4, 1)),
        ('SPAN', (0, 2), (4, 2)),
        ('BACKGROUND', (0, 3), (-1, 3), colors.darkolivegreen),
        ('TEXTCOLOR', (0, 3), (-1, 3), colors.whitesmoke),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story.append(t3)

    doc.build(story)
    print(f"[OK] PDF de muestra generado en: {output_path}")


def create_sample_excel(output_path: str):
    """Crea un archivo Excel con encabezados desplazados, celdas de inicio específicas y tablas con nombre oficial."""
    import openpyxl
    from openpyxl.worksheet.table import Table as OpxTable, TableStyleInfo
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    wb = openpyxl.Workbook()
    
    # --- HOJA 1: Inventario (Con Tabla Oficial de Excel 'TablaStock' en A4:F10) ---
    ws1 = wb.active
    ws1.title = "Inventario"
    ws1['A1'] = "SISTEMA DE GESTION DE BODEGA"
    ws1['A2'] = "Reporte de Stock Valorizado - Corte: 2026-08-20"
    ws1['A3'] = "Filtro: Todos los almacenes activos"
    
    headers_inv = ["ID_Item", "Nombre_Articulo", "Bodega", "Stock", "Costo_Unitario", "Valor_Total"]
    for col_i, h in enumerate(headers_inv, start=1):
        ws1.cell(row=4, column=col_i, value=h)
    
    datos_inv = [
        ["ART-001", "Tuerca Hexagonal M8 (Pack 100)", "Bodega Central", 150, 4.50, 675.00],
        ["ART-002", "Perno Acero Inox 2 pulg", "Bodega Central", 320, 1.20, 384.00],
        ["ART-003", "Disco Corte Metal 4 1/2", "Bodega Norte", 85, 3.80, 323.00],
        ["ART-004", "Guantes Nitrilo Talla L", "Bodega Sur", 200, 2.10, 420.00],
        ["ART-005", "Soldadura 6011 1/8 (Kg)", "Bodega Central", 45, 12.50, 562.50],
        ["ART-006", "Cinta Aisladora 3M Negra", "Bodega Norte", 110, 1.95, 214.50],
    ]
    for r_i, row in enumerate(datos_inv, start=5):
        for c_i, val in enumerate(row, start=1):
            ws1.cell(row=r_i, column=c_i, value=val)

    # Agregar como Named Table oficial de Excel
    tab_stock = OpxTable(displayName="TablaStock", ref="A4:F10")
    ws1.add_table(tab_stock)

    # --- HOJA 2: Proveedores (Encabezado en Fila 3) ---
    ws2 = wb.create_sheet(title="Proveedores")
    ws2['A1'] = "EVALUACION DE PROVEEDORES 2026"
    ws2['A2'] = "Area de Adquisiciones y Abastecimiento"
    headers_prov = ["Proveedor", "Calificacion", "Cumplimiento_Plazos", "Condicion_Pago"]
    for c_i, h in enumerate(headers_prov, start=1):
        ws2.cell(row=3, column=c_i, value=h)
    
    datos_prov = [
        ["Ferreteria Industrial SpA", 9.2, 98.5, "30 dias"],
        ["Aceros y Metales del Pacifico", 8.7, 92.0, "60 dias"],
        ["Componentes Electronicos SA", 9.5, 99.0, "Contado"],
        ["Suministros Mineros Ltda", 7.8, 85.0, "45 dias"],
    ]
    for r_i, row in enumerate(datos_prov, start=4):
        for c_i, val in enumerate(row, start=1):
            ws2.cell(row=r_i, column=c_i, value=val)

    # --- HOJA 3: Despacho (Tabla parte exactamente en la celda C4) ---
    ws3 = wb.create_sheet(title="Despacho")
    ws3['A1'] = "LOGISTICA DE ENVIO"
    ws3['C2'] = "CONTROL DE RUTAS REGIONALES"
    # Tabla en C4:F8
    headers_desp = ["Codigo_Ruta", "Chofer", "Vehiculo", "Estado"]
    for c_i, h in enumerate(headers_desp, start=3): # Col C es 3
        ws3.cell(row=4, column=c_i, value=h)
    
    datos_desp = [
        ["RUT-01", "Pedro Gomez", "Camioneta Ford", "En Transito"],
        ["RUT-02", "Juan Perez", "Furgon Peugeot", "Entregado"],
        ["RUT-03", "Mario Soto", "Camion Mercedes", "En Carga"],
        ["RUT-04", "Luis Varas", "Furgon Hyundai", "Pendiente"],
    ]
    for r_i, row in enumerate(datos_desp, start=5):
        for c_i, val in enumerate(row, start=3):
            ws3.cell(row=r_i, column=c_i, value=val)

    wb.save(output_path)
    print(f"[OK] Excel de muestra generado en: {output_path}")


def create_sample_sqlite(output_path: str):
    """Crea una base de datos SQLite de muestra con varias tablas y vistas."""
    import sqlite3
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if os.path.exists(output_path):
        os.remove(output_path)

    with sqlite3.connect(output_path) as conn:
        cursor = conn.cursor()
        
        # Tabla Clientes
        cursor.execute("""
            CREATE TABLE clientes (
                id_cliente INTEGER PRIMARY KEY,
                nombre TEXT NOT NULL,
                ciudad TEXT,
                segmento TEXT,
                credito_max REAL
            )
        """)
        clientes_data = [
            (1, "Distribuidora Los Andes", "Santiago", "Corporativo", 50000.0),
            (2, "Comercial del Sur SpA", "Concepcion", "PYME", 15000.0),
            (3, "Servicios Mineros Antofagasta", "Antofagasta", "Corporativo", 120000.0),
            (4, "Importadora Valparaiso", "Valparaiso", "PYME", 25000.0),
            (5, "Ferreteria Central", "Rancagua", "Minorista", 8000.0),
        ]
        cursor.executemany("INSERT INTO clientes VALUES (?, ?, ?, ?, ?)", clientes_data)

        # Tabla Ventas
        cursor.execute("""
            CREATE TABLE ventas (
                id_venta INTEGER PRIMARY KEY,
                id_cliente INTEGER,
                fecha TEXT,
                monto_neto REAL,
                estado TEXT
            )
        """)
        ventas_data = [
            (101, 1, "2026-08-01", 12500.0, "Pagado"),
            (102, 3, "2026-08-03", 45000.0, "Pagado"),
            (103, 2, "2026-08-05", 8200.0, "Pendiente"),
            (104, 1, "2026-08-10", 16800.0, "Pagado"),
            (105, 4, "2026-08-12", 9400.0, "Pagado"),
            (106, 5, "2026-08-15", 3100.0, "Pendiente"),
        ]
        cursor.executemany("INSERT INTO ventas VALUES (?, ?, ?, ?, ?)", ventas_data)

        # Vista Resumen
        cursor.execute("""
            CREATE VIEW vista_resumen_clientes AS
            SELECT c.nombre, c.ciudad, SUM(v.monto_neto) AS total_compras, COUNT(v.id_venta) AS cantidad_pedidos
            FROM clientes c
            JOIN ventas v ON c.id_cliente = v.id_cliente
            GROUP BY c.id_cliente
        """)
        conn.commit()

    print(f"[OK] SQLite de muestra generado en: {output_path}")


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    samples_dir = os.path.join(base_dir, "samples")
    
    pdf_path = os.path.join(samples_dir, "ejemplo_facturas.pdf")
    excel_path = os.path.join(samples_dir, "ejemplo_inventario.xlsx")
    sqlite_path = os.path.join(samples_dir, "ejemplo_empresa.db")

    create_sample_pdf(pdf_path)
    create_sample_excel(excel_path)
    create_sample_sqlite(sqlite_path)
