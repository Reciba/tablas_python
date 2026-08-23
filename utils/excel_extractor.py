"""
Módulo para extracción de tablas desde archivos Excel (.xlsx, .xls, .xlsm) y CSV usando xlwings.
Soporta:
1. Reconocimiento por nombres de tablas oficiales de Excel (ListObjects / Named Tables).
2. Extracción por celda de inicio donde parte la tabla (ej: celda_inicio="C4" o "B3").
3. Extracción por rango exacto (ej: rango="C4:F20").
4. Archivos ABIERTOS o CERRADOS en Microsoft Excel.
5. Rutas relativas o absolutas.
"""

from typing import List, Any, Optional, Union, Dict, Tuple
from dataclasses import dataclass
import os
import re
import pandas as pd

# Intentar importar xlwings
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

try:
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class RawExcelTableInfo:
    """Información y contenido crudo de una tabla u hoja detectada en Excel/CSV."""
    table_id: int
    source_file: str
    sheet_name: str
    raw_data: List[List[Any]]
    num_rows: int
    num_cols: int
    table_name: Optional[str] = None
    start_cell: Optional[str] = None
    range_address: Optional[str] = None

    def get_preview(self, max_rows: int = 5) -> List[List[Any]]:
        """Retorna las primeras filas crudas para inspección visual."""
        return self.raw_data[:max_rows]

from .file_utils import resolve_file_path


class ExcelTableExtractor:
    """
    Extractor de tablas para Excel basado en xlwings con soporte para:
    - Tablas con nombre oficial de Excel (ListObjects).
    - Celdas de inicio (ej: C4) y rangos (ej: B3:F15).
    - Archivos abiertos o cerrados.
    """

    def __init__(
        self,
        file_path: str,
        archivo_abierto: Optional[bool] = None,
        preferir_xlwings: bool = True
    ):
        self.archivo_abierto = archivo_abierto
        self.preferir_xlwings = preferir_xlwings and HAS_XLWINGS
        self.book_name = os.path.basename(file_path) if file_path else ""
        self.is_open_in_memory = False

        # 1. Si el usuario indicó archivo_abierto=True o None, buscar primero en libros abiertos de xlwings
        if self.preferir_xlwings and archivo_abierto is not False:
            matched_book = self._find_open_workbook_in_excel(file_path)
            if matched_book:
                self.book_name = matched_book.name
                self.file_path = getattr(matched_book, 'fullname', matched_book.name)
                self.ext = os.path.splitext(matched_book.name)[1].lower() or ".xlsx"
                self.is_open_in_memory = True
                return

        # Si forzó archivo_abierto=True pero no se encontró libro abierto en Excel
        if archivo_abierto is True and self.preferir_xlwings:
            abiertos = [b.name for b in xw.books] if HAS_XLWINGS and len(xw.books) > 0 else []
            raise FileNotFoundError(
                f"No se encontró ningún libro abierto en Excel con el nombre '{file_path}'. "
                f"Libros actualmente abiertos en Excel: {abiertos if abiertos else 'Ninguno (Excel no tiene libros abiertos)'}"
            )

        # 2. Si es un archivo cerrado o en disco, resolver la ruta física
        self.file_path = resolve_file_path(file_path)
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"No se encontró el archivo Excel en: {self.file_path}")
        
        self.ext = os.path.splitext(self.file_path)[1].lower()

    def _find_open_workbook_in_excel(self, target: str) -> Optional[Any]:
        """Busca y retorna el objeto Book de xlwings si está abierto en memoria."""
        if not HAS_XLWINGS:
            return None
        try:
            if not target or target.lower() in ["", "activo", "active", "libro_activo", "hoja_activa"]:
                if len(xw.books) > 0:
                    return xw.books.active
                return None

            clean_target = os.path.basename(target).strip().lower()
            name_no_ext = os.path.splitext(clean_target)[0]

            for book in xw.books:
                b_name = book.name.lower()
                b_name_no_ext = os.path.splitext(b_name)[0]
                b_fullname = getattr(book, 'fullname', '').lower()

                # Comparar con nombre completo, sin extensión o fullname
                if (clean_target == b_name or
                    name_no_ext == b_name_no_ext or
                    (b_fullname and clean_target == os.path.basename(b_fullname).lower())):
                    return book
        except Exception:
            return None
        return None

    def _is_workbook_open_in_excel(self, file_name: str) -> bool:
        """Comprueba si el libro ya está abierto en alguna sesión activa de Excel."""
        return self._find_open_workbook_in_excel(file_name) is not None

    def extract_by_cell_or_range(
        self,
        celda_inicio: Optional[str] = None,
        rango: Optional[str] = None,
        hoja: Optional[Union[str, int]] = None
    ) -> List[List[Any]]:
        """
        Extrae datos a partir de una celda de inicio (ej: 'C4') o un rango específico (ej: 'C4:G20').
        """
        if self.preferir_xlwings:
            try:
                return self._extract_range_xlwings(celda_inicio=celda_inicio, rango=rango, hoja=hoja)
            except Exception:
                return self._extract_range_openpyxl(celda_inicio=celda_inicio, rango=rango, hoja=hoja)
        else:
            return self._extract_range_openpyxl(celda_inicio=celda_inicio, rango=rango, hoja=hoja)

    def _extract_range_xlwings(
        self,
        celda_inicio: Optional[str] = None,
        rango: Optional[str] = None,
        hoja: Optional[Union[str, int]] = None
    ) -> List[List[Any]]:
        file_name = os.path.basename(self.file_path)
        is_already_open = self._is_workbook_open_in_excel(file_name)
        should_connect = self.archivo_abierto is True or (self.archivo_abierto is None and is_already_open)

        app = None
        book = None
        needs_close = False

        try:
            if should_connect:
                try:
                    book = xw.books[file_name]
                except Exception:
                    book = xw.Book(self.file_path)
            else:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False
                book = app.books.open(self.file_path, read_only=True)
                needs_close = True

            # Seleccionar hoja
            if hoja is None:
                sht = book.sheets[0]
            elif isinstance(hoja, int):
                sht = book.sheets[hoja - 1]
            else:
                sht = book.sheets[hoja]

            # Leer por rango o por celda_inicio
            if rango:
                val = sht.range(rango).value
            elif celda_inicio:
                # Expandir tabla a partir de la celda de inicio
                clean_cell = celda_inicio.replace("$", "").upper()
                val = sht.range(clean_cell).expand('table').value
            else:
                val = sht.used_range.value

            if val is None:
                return []
            if not isinstance(val, list):
                return [[val]]
            if val and not isinstance(val[0], list):
                return [val]
            return val

        finally:
            if needs_close:
                if book:
                    try: book.close()
                    except Exception: pass
                if app:
                    try: app.quit()
                    except Exception: pass

    def _extract_range_openpyxl(
        self,
        celda_inicio: Optional[str] = None,
        rango: Optional[str] = None,
        hoja: Optional[Union[str, int]] = None
    ) -> List[List[Any]]:
        """Fallback con openpyxl para extraer por celda de inicio o rango."""
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        if hoja is None:
            ws = wb.active
        elif isinstance(hoja, int):
            ws = wb.worksheets[hoja - 1]
        else:
            ws = wb[hoja]

        if rango:
            cells = ws[rango]
            if isinstance(cells, tuple):
                if isinstance(cells[0], tuple):
                    return [[c.value for c in row] for row in cells]
                return [[c.value for c in cells]]
            return [[cells.value]]

        if celda_inicio:
            row_idx, col_idx = coordinate_to_tuple(celda_inicio.replace("$", "").upper())
            # Leer desde row_idx, col_idx hasta el final de la región de datos
            rows = []
            for r in range(row_idx, ws.max_row + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(col_idx, ws.max_column + 1)]
                # Detener si la fila está completamente vacía
                if all(v is None for v in row_vals):
                    break
                rows.append(row_vals)
            
            # Recortar columnas vacías al final
            if rows:
                max_valid_col = 0
                for r in rows:
                    for i in reversed(range(len(r))):
                        if r[i] is not None:
                            max_valid_col = max(max_valid_col, i + 1)
                            break
                if max_valid_col > 0:
                    rows = [r[:max_valid_col] for r in rows]
            return rows

        # Toda la hoja
        return [[c.value for c in row] for row in ws.iter_rows()]

    def extract_all_tables(
        self,
        sheets: Optional[List[Union[str, int]]] = None,
        split_blank_blocks: bool = False
    ) -> List[RawExcelTableInfo]:
        """
        Extrae todas las tablas u hojas del archivo Excel.
        Detecta automáticamente si existen tablas con nombre oficial (ListObjects) en las hojas.
        """
        if self.ext == '.csv':
            df_raw = pd.read_csv(self.file_path, header=None, dtype=object)
            raw_list = df_raw.values.tolist()
            info = RawExcelTableInfo(
                table_id=1,
                source_file=self.file_path,
                sheet_name="CSV",
                table_name="CSV",
                raw_data=raw_list,
                num_rows=len(raw_list),
                num_cols=len(raw_list[0]) if raw_list else 0
            )
            return [info]

        if self.preferir_xlwings:
            try:
                return self._extract_all_xlwings(sheets=sheets)
            except Exception:
                return self._extract_all_openpyxl(sheets=sheets, split_blank_blocks=split_blank_blocks)
        else:
            return self._extract_all_openpyxl(sheets=sheets, split_blank_blocks=split_blank_blocks)

    def _extract_all_xlwings(
        self,
        sheets: Optional[List[Union[str, int]]] = None
    ) -> List[RawExcelTableInfo]:
        file_name = os.path.basename(self.file_path)
        is_already_open = self._is_workbook_open_in_excel(file_name)
        should_connect = self.archivo_abierto is True or (self.archivo_abierto is None and is_already_open)

        tables_found: List[RawExcelTableInfo] = []
        app = None
        book = None
        needs_close = False

        try:
            if should_connect:
                try:
                    book = xw.books[file_name]
                except Exception:
                    book = xw.Book(self.file_path)
            else:
                app = xw.App(visible=False, add_book=False)
                app.display_alerts = False
                app.screen_updating = False
                book = app.books.open(self.file_path, read_only=True)
                needs_close = True

            all_sheet_names = [s.name for s in book.sheets]
            target_sheets = []
            if sheets is None:
                target_sheets = all_sheet_names
            else:
                for s in sheets:
                    if isinstance(s, int) and 1 <= s <= len(all_sheet_names):
                        target_sheets.append(all_sheet_names[s - 1])
                    elif str(s) in all_sheet_names:
                        target_sheets.append(str(s))

            global_id = 1
            for sheet_name in target_sheets:
                sht = book.sheets[sheet_name]
                
                # 1. Comprobar si la hoja tiene tablas oficiales de Excel (ListObjects)
                named_tables = list(sht.tables)
                if named_tables:
                    for nt in named_tables:
                        t_vals = nt.range.value
                        if not isinstance(t_vals, list):
                            t_vals = [[t_vals]]
                        elif t_vals and not isinstance(t_vals[0], list):
                            t_vals = [t_vals]

                        addr = nt.range.address
                        start_c = addr.split(":")[0].replace("$", "") if ":" in addr else addr.replace("$", "")

                        info = RawExcelTableInfo(
                            table_id=global_id,
                            source_file=self.file_path,
                            sheet_name=sheet_name,
                            table_name=nt.name,
                            start_cell=start_c,
                            range_address=addr,
                            raw_data=t_vals,
                            num_rows=len(t_vals),
                            num_cols=max((len(r) for r in t_vals if isinstance(r, list)), default=0)
                        )
                        tables_found.append(info)
                        global_id += 1
                else:
                    # Hoja estándar
                    used_range = sht.used_range
                    raw_values = used_range.value
                    if raw_values is None:
                        continue
                    if not isinstance(raw_values, list):
                        raw_matrix = [[raw_values]]
                    elif raw_values and not isinstance(raw_values[0], list):
                        raw_matrix = [raw_values]
                    else:
                        raw_matrix = raw_values

                    addr = used_range.address
                    start_c = addr.split(":")[0].replace("$", "") if ":" in addr else "A1"

                    info = RawExcelTableInfo(
                        table_id=global_id,
                        source_file=self.file_path,
                        sheet_name=sheet_name,
                        table_name=sheet_name,
                        start_cell=start_c,
                        range_address=addr,
                        raw_data=raw_matrix,
                        num_rows=len(raw_matrix),
                        num_cols=max((len(r) for r in raw_matrix if isinstance(r, list)), default=0)
                    )
                    tables_found.append(info)
                    global_id += 1

            return tables_found

        finally:
            if needs_close:
                if book:
                    try: book.close()
                    except Exception: pass
                if app:
                    try: app.quit()
                    except Exception: pass

    def _extract_all_openpyxl(
        self,
        sheets: Optional[List[Union[str, int]]] = None,
        split_blank_blocks: bool = False
    ) -> List[RawExcelTableInfo]:
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        all_sheet_names = wb.sheetnames
        
        target_sheets = []
        if sheets is None:
            target_sheets = all_sheet_names
        else:
            for s in sheets:
                if isinstance(s, int) and 1 <= s <= len(all_sheet_names):
                    target_sheets.append(all_sheet_names[s - 1])
                elif str(s) in all_sheet_names:
                    target_sheets.append(str(s))

        tables_found: List[RawExcelTableInfo] = []
        global_id = 1

        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            
            # Tablas oficiales openpyxl
            if hasattr(ws, 'tables') and ws.tables:
                for t_name, t_obj in ws.tables.items():
                    # t_obj puede ser el objeto Table o tener ref
                    ref = getattr(t_obj, 'ref', str(t_obj))
                    cells = ws[ref]
                    if isinstance(cells, tuple):
                        if isinstance(cells[0], tuple):
                            raw_matrix = [[c.value for c in row] for row in cells]
                        else:
                            raw_matrix = [[c.value for c in cells]]
                    else:
                        raw_matrix = [[cells.value]]

                    start_c = ref.split(":")[0] if ":" in ref else ref

                    info = RawExcelTableInfo(
                        table_id=global_id,
                        source_file=self.file_path,
                        sheet_name=sheet_name,
                        table_name=t_name,
                        start_cell=start_c,
                        range_address=ref,
                        raw_data=raw_matrix,
                        num_rows=len(raw_matrix),
                        num_cols=max((len(r) for r in raw_matrix if isinstance(r, list)), default=0)
                    )
                    tables_found.append(info)
                    global_id += 1
            else:
                # Leer hoja completa
                raw_matrix = [[c.value for c in row] for row in ws.iter_rows()]
                if not raw_matrix:
                    continue
                info = RawExcelTableInfo(
                    table_id=global_id,
                    source_file=self.file_path,
                    sheet_name=sheet_name,
                    table_name=sheet_name,
                    start_cell="A1",
                    range_address=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                    raw_data=raw_matrix,
                    num_rows=len(raw_matrix),
                    num_cols=max((len(r) for r in raw_matrix if isinstance(r, list)), default=0)
                )
                tables_found.append(info)
                global_id += 1

        return tables_found
