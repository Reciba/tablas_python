"""
Módulo para escribir y actualizar datos en archivos Excel usando xlwings.
Permite pegar DataFrames directamente en celdas específicas (ej: 'B5') de libros
abiertos en pantalla (en vivo) o cerrados en segundo plano, sin dañar formatos ni fórmulas existentes.
"""

from typing import Optional, Union, Any
import os
import pandas as pd

try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False

try:
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .file_utils import resolve_file_path


def _find_open_workbook_in_excel(target: str) -> Optional[Any]:
    """Busca y retorna el objeto Book de xlwings si está abierto en memoria."""
    if not HAS_XLWINGS:
        return None
    try:
        if not target or target.lower() in ["", "activo", "active", "libro_activo"]:
            if len(xw.books) > 0:
                return xw.books.active
            return None

        clean_target = os.path.basename(target).strip().lower()
        name_no_ext = os.path.splitext(clean_target)[0]

        for book in xw.books:
            b_name = book.name.lower()
            b_name_no_ext = os.path.splitext(b_name)[0]
            b_fullname = getattr(book, 'fullname', '').lower()

            if (clean_target == b_name or
                name_no_ext == b_name_no_ext or
                (b_fullname and clean_target == os.path.basename(b_fullname).lower())):
                return book
    except Exception:
        return None
    return None


def _is_workbook_open_in_excel(file_path: str, file_name: str) -> bool:
    """Comprueba si el libro ya está abierto en Excel activo."""
    return _find_open_workbook_in_excel(file_name) is not None or _find_open_workbook_in_excel(file_path) is not None


def escribir_en_excel(
    df: pd.DataFrame,
    archivo_excel: str,
    hoja: Union[str, int] = 1,
    celda_inicio: str = "A1",
    incluir_encabezados: bool = True,
    incluir_indice: bool = False,
    guardar: bool = True,
    archivo_abierto: Optional[bool] = None,
    crear_si_no_existe: bool = True
) -> str:
    """
    Escribe un DataFrame directamente en una celda específica de un libro de Excel.
    """
    # 1. Si está abierto en memoria o archivo_abierto=True, usar directamente xlwings
    if HAS_XLWINGS and archivo_abierto is not False:
        matched_book = _find_open_workbook_in_excel(archivo_excel)
        if matched_book:
            book = matched_book
            # Seleccionar o crear hoja
            sheet_names = [s.name for s in book.sheets]
            if isinstance(hoja, int):
                if 1 <= hoja <= len(book.sheets):
                    sht = book.sheets[hoja - 1]
                else:
                    sht = book.sheets.add(f"Hoja{hoja}")
            else:
                if str(hoja) in sheet_names:
                    sht = book.sheets[str(hoja)]
                else:
                    sht = book.sheets.add(str(hoja))

            clean_cell = celda_inicio.replace("$", "").upper()
            sht.range(clean_cell).options(
                index=incluir_indice,
                header=incluir_encabezados
            ).value = df

            if guardar:
                try: book.save()
                except Exception: pass

            return getattr(book, 'fullname', book.name)

    if archivo_abierto is True and HAS_XLWINGS:
        abiertos = [b.name for b in xw.books] if len(xw.books) > 0 else []
        raise FileNotFoundError(
            f"No se encontró ningún libro abierto en Excel con el nombre '{archivo_excel}'. "
            f"Libros actualmente abiertos en Excel: {abiertos if abiertos else 'Ninguno (Excel no tiene libros abiertos)'}"
        )

    # 2. Si es archivo cerrado en disco
    try:
        resolved_path = resolve_file_path(archivo_excel)
    except Exception:
        resolved_path = os.path.abspath(archivo_excel)

    if not os.path.exists(resolved_path):
        if crear_si_no_existe:
            os.makedirs(os.path.dirname(resolved_path), exist_ok=True)
            df_init = pd.DataFrame()
            with pd.ExcelWriter(resolved_path, engine='openpyxl') as writer:
                sheet_title = hoja if isinstance(hoja, str) else "Hoja1"
                df_init.to_excel(writer, sheet_name=sheet_title)
        else:
            raise FileNotFoundError(f"No existe el archivo Excel: {resolved_path}")

    file_name = os.path.basename(resolved_path)

    # Intentar con xlwings si está disponible
    if HAS_XLWINGS:
        app = None
        book = None
        try:
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False
            book = app.books.open(resolved_path)

            sheet_names = [s.name for s in book.sheets]
            if isinstance(hoja, int):
                if 1 <= hoja <= len(book.sheets):
                    sht = book.sheets[hoja - 1]
                else:
                    sht = book.sheets.add(f"Hoja{hoja}")
            else:
                if str(hoja) in sheet_names:
                    sht = book.sheets[str(hoja)]
                else:
                    sht = book.sheets.add(str(hoja))

            clean_cell = celda_inicio.replace("$", "").upper()
            sht.range(clean_cell).options(
                index=incluir_indice,
                header=incluir_encabezados
            ).value = df

            if guardar:
                book.save()

            return resolved_path
        finally:
            if book:
                try: book.close()
                except Exception: pass
            if app:
                try: app.quit()
                except Exception: pass

        return resolved_path

    # Fallback con openpyxl si xlwings no pudo ejecutarse
    if HAS_OPENPYXL:
        wb = openpyxl.load_workbook(resolved_path)
        sheet_title = hoja if isinstance(hoja, str) else (f"Hoja{hoja}" if isinstance(hoja, int) and hoja > len(wb.sheetnames) else wb.sheetnames[hoja-1])
        if sheet_title in wb.sheetnames:
            ws = wb[sheet_title]
        else:
            ws = wb.create_sheet(title=sheet_title)

        start_row, start_col = coordinate_to_tuple(celda_inicio.replace("$", "").upper())

        current_r = start_row
        if incluir_encabezados:
            for c_idx, col_name in enumerate(df.columns, start=start_col):
                ws.cell(row=current_r, column=c_idx, value=str(col_name))
            current_r += 1

        for _, row in df.iterrows():
            for c_idx, val in enumerate(row.values, start=start_col):
                cell_val = None if pd.isna(val) else val
                ws.cell(row=current_r, column=c_idx, value=cell_val)
            current_r += 1

        if guardar:
            wb.save(resolved_path)
        return resolved_path

    raise RuntimeError("Se requiere xlwings u openpyxl para escribir en Excel.")
